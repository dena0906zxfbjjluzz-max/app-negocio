"""Generación de Excel (cuaderno) y PDF de resumen semanal."""

from __future__ import annotations

import io
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

# Colores tipo cuaderno / corporativo (inspirado en validador, tono campo)
AZUL = "1F4E78"
AZUL_CLARO = "D6E3F0"
VERDE = "1B5E3B"
VERDE_CLARO = "E8F5E9"
AMARILLO = "FFF3CD"
ROJO = "F8D7DA"
GRIS = "F2F2F2"
BLANCO = "FFFFFF"


def _borde():
    s = Side(style="thin", color="B0B0B0")
    return Border(left=s, right=s, top=s, bottom=s)


def _auto_ancho(ws, n_cols: int, last_row: int) -> None:
    for c in range(1, n_cols + 1):
        letter = get_column_letter(c)
        max_len = 10
        for r in range(1, last_row + 1):
            val = ws.cell(row=r, column=c).value
            if val is None:
                continue
            for parte in str(val).splitlines() or [""]:
                max_len = max(max_len, len(parte))
        ws.column_dimensions[letter].width = min(max(max_len + 2, 11), 36)


def _pintar_titulo(ws, titulo: str, n_cols: int, fill_hex: str = AZUL) -> None:
    if n_cols > 1:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    cell = ws.cell(row=1, column=1, value=titulo)
    cell.font = Font(name="Calibri", bold=True, color=BLANCO, size=14)
    cell.fill = PatternFill("solid", fgColor=fill_hex)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 28
    for c in range(1, n_cols + 1):
        ws.cell(row=1, column=c).fill = PatternFill("solid", fgColor=fill_hex)
        ws.cell(row=1, column=c).border = _borde()


def _escribir_headers(ws, headers: list[str], fila: int = 2) -> None:
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=fila, column=c, value=h)
        cell.font = Font(name="Calibri", bold=True, color=BLANCO, size=11)
        cell.fill = PatternFill("solid", fgColor=AZUL)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _borde()
    ws.row_dimensions[fila].height = 22


def _fill_estado(estado: str) -> PatternFill | None:
    e = (estado or "").lower()
    if "fiado" in e or "debe" in e:
        return PatternFill("solid", fgColor=ROJO)
    if "transfer" in e:
        return PatternFill("solid", fgColor=AMARILLO)
    if "efectivo" in e or "pagado" in e:
        return PatternFill("solid", fgColor=VERDE_CLARO)
    return None


def ventas_a_dataframe(ventas: list[dict]) -> pd.DataFrame:
    if not ventas:
        return pd.DataFrame(
            columns=["Día", "Cliente", "Sacos", "Precio S/", "Monto S/", "Estado"]
        )
    rows = []
    for v in ventas:
        sacos = int(v.get("sacos", 0))
        precio = float(v.get("precio", 0))
        rows.append(
            {
                "Día": v.get("dia", ""),
                "Cliente": v.get("cliente", ""),
                "Sacos": sacos,
                "Precio S/": round(precio, 2),
                "Monto S/": round(sacos * precio, 2),
                "Estado": v.get("estado", ""),
            }
        )
    df = pd.DataFrame(rows)
    # Orden por día de la semana si es posible
    orden = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado", "Domingo"]
    df["_ord"] = df["Día"].map({d: i for i, d in enumerate(orden)}).fillna(99)
    df = df.sort_values(["_ord", "Cliente"]).drop(columns=["_ord"])
    return df.reset_index(drop=True)


def _hoja_tabla(
    wb: Workbook,
    nombre: str,
    titulo: str,
    df: pd.DataFrame,
    *,
    resaltar_estado: bool = True,
    fill_titulo: str = AZUL,
) -> None:
    ws = wb.create_sheet(title=nombre[:31])
    if df is None or df.empty:
        headers = ["Aviso"]
        _pintar_titulo(ws, titulo, 1, fill_titulo)
        _escribir_headers(ws, headers)
        ws.cell(row=3, column=1, value="Sin datos en este periodo")
        ws.cell(row=3, column=1).border = _borde()
        _auto_ancho(ws, 1, 3)
        return

    headers = list(df.columns)
    n_cols = len(headers)
    _pintar_titulo(ws, titulo, n_cols, fill_titulo)
    _escribir_headers(ws, headers)

    estado_col = None
    if resaltar_estado and "Estado" in headers:
        estado_col = headers.index("Estado") + 1

    for r_idx, row in enumerate(df.itertuples(index=False), start=3):
        zebra = r_idx % 2 == 0
        for c_idx, valor in enumerate(row, start=1):
            if pd.isna(valor):
                valor = ""
            cell = ws.cell(row=r_idx, column=c_idx, value=valor)
            cell.font = Font(name="Calibri", size=10)
            cell.border = _borde()
            cell.alignment = Alignment(
                horizontal="center" if c_idx > 2 else "left",
                vertical="center",
            )
            # Estado coloreado
            if estado_col and c_idx == estado_col:
                fill = _fill_estado(str(valor))
                if fill:
                    cell.fill = fill
                    continue
            if zebra:
                cell.fill = PatternFill("solid", fgColor=AZUL_CLARO)

    # Fila de totales si hay montos
    last = 2 + len(df)
    if "Sacos" in headers or "Monto S/" in headers:
        last += 1
        ws.cell(row=last, column=1, value="TOTAL")
        for c in range(1, n_cols + 1):
            cell = ws.cell(row=last, column=c)
            cell.font = Font(name="Calibri", bold=True, size=11, color=BLANCO)
            cell.fill = PatternFill("solid", fgColor=VERDE)
            cell.border = _borde()
        if "Sacos" in headers:
            ci = headers.index("Sacos") + 1
            ws.cell(row=last, column=ci, value=int(df["Sacos"].sum()))
        if "Monto S/" in headers:
            ci = headers.index("Monto S/") + 1
            ws.cell(row=last, column=ci, value=round(float(df["Monto S/"].sum()), 2))

    _auto_ancho(ws, n_cols, last)
    ws.freeze_panes = "A3"
    ws.print_title_rows = "1:2"


def _hoja_portada(
    wb: Workbook,
    semana: str,
    df: pd.DataFrame,
    faltantes: list[str],
) -> None:
    ws = wb.create_sheet(title="Resumen", index=0)
    _pintar_titulo(ws, f"CUADERNO DE DESPACHOS — {semana}", 4, VERDE)

    metrics = [
        ("Fecha de emisión", datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("Semana", semana),
        ("Total despachos", str(len(df))),
        ("Total sacos", str(int(df["Sacos"].sum()) if not df.empty else 0)),
        (
            "Facturación S/",
            f"{float(df['Monto S/'].sum()):.2f}" if not df.empty and "Monto S/" in df.columns else "0.00",
        ),
    ]
    if not df.empty and "Estado" in df.columns:
        fiados = df[df["Estado"].str.contains("Fiado|Debe", case=False, na=False)]
        metrics.append(
            (
                "Por cobrar (fiados) S/",
                f"{float(fiados['Monto S/'].sum()):.2f}" if not fiados.empty else "0.00",
            )
        )

    _escribir_headers(ws, ["Indicador", "Valor", "", ""], fila=2)
    for i, (k, v) in enumerate(metrics, start=3):
        ws.cell(row=i, column=1, value=k).border = _borde()
        ws.cell(row=i, column=2, value=v).border = _borde()
        ws.cell(row=i, column=1).font = Font(name="Calibri", bold=True, size=11)
        ws.cell(row=i, column=2).fill = PatternFill("solid", fgColor=GRIS)
        for c in (3, 4):
            ws.cell(row=i, column=c).border = _borde()

    row = 3 + len(metrics) + 1
    ws.cell(row=row, column=1, value="Pollerías sin pedido esta semana")
    ws.cell(row=row, column=1).font = Font(name="Calibri", bold=True, size=12, color=AZUL)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)

    row += 1
    if faltantes:
        for c_name in faltantes:
            ws.cell(row=row, column=1, value=c_name).fill = PatternFill("solid", fgColor=ROJO)
            ws.cell(row=row, column=1).border = _borde()
            row += 1
    else:
        ws.cell(row=row, column=1, value="Todas las de la lista tienen despacho").fill = (
            PatternFill("solid", fgColor=VERDE_CLARO)
        )
        ws.cell(row=row, column=1).border = _borde()
        row += 1

    _auto_ancho(ws, 4, row)
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 22


def generar_excel_cuaderno(
    ventas: list[dict],
    semana: str,
    lista_clientes: list[str] | None = None,
) -> io.BytesIO:
    """
    Excel multi-hoja estilo cuaderno:
    - Resumen (portada con totales)
    - Cuaderno (todo el detalle)
    - Una hoja por día con datos
    - Por cobrar (fiados)
    """
    df = ventas_a_dataframe(ventas)
    lista_clientes = lista_clientes or []
    con_pedido = set(df["Cliente"].unique()) if not df.empty else set()
    faltantes = [c for c in lista_clientes if c not in con_pedido]

    wb = Workbook()
    # quitar default
    default = wb.active
    wb.remove(default)

    _hoja_portada(wb, semana, df, faltantes)
    _hoja_tabla(
        wb,
        "Cuaderno",
        f"Detalle completo · {semana}",
        df,
        fill_titulo=AZUL,
    )

    orden_dias = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado", "Domingo"]
    if not df.empty:
        for dia in orden_dias:
            parte = df[df["Día"] == dia]
            if parte.empty:
                continue
            corto = {"Lunes": "Lun", "Martes": "Mar", "Miércoles": "Mie", "Jueves": "Jue",
                     "Viernes": "Vie", "Sábado": "Sab", "Domingo": "Dom"}.get(dia, dia[:3])
            _hoja_tabla(
                wb,
                corto,
                f"Hoja del {dia} · {semana}",
                parte.drop(columns=["Día"]) if "Día" in parte.columns else parte,
            )

    if not df.empty:
        fiados = df[df["Estado"].astype(str).str.contains("Fiado|Debe", case=False, na=False)].copy()
    else:
        fiados = df.copy()
    _hoja_tabla(
        wb,
        "Por cobrar",
        f"Cuentas por cobrar · {semana}",
        fiados,
        fill_titulo="8B0000",
    )

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def generar_pdf_resumen(
    ventas: list[dict],
    semana: str,
    lista_clientes: list[str] | None = None,
    negocio: str = "Control de Papas y Pollerías",
) -> io.BytesIO:
    """PDF ejecutivo similar en estilo al del validador (tabla + indicadores)."""
    df = ventas_a_dataframe(ventas)
    lista_clientes = lista_clientes or []
    con_pedido = set(df["Cliente"].unique()) if not df.empty else set()
    faltantes = [c for c in lista_clientes if c not in con_pedido]

    total_sacos = int(df["Sacos"].sum()) if not df.empty else 0
    total_monto = float(df["Monto S/"].sum()) if not df.empty else 0.0
    n_despachos = len(df)
    if not df.empty:
        fiados = df[df["Estado"].astype(str).str.contains("Fiado|Debe", case=False, na=False)]
        monto_fiado = float(fiados["Monto S/"].sum()) if not fiados.empty else 0.0
        n_fiados = len(fiados)
    else:
        monto_fiado = 0.0
        n_fiados = 0

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        rightMargin=30,
        leftMargin=30,
        topMargin=30,
        bottomMargin=30,
    )
    styles = getSampleStyleSheet()
    titulo_estilo = ParagraphStyle(
        "TituloCuaderno",
        parent=styles["Heading1"],
        fontSize=15,
        textColor=colors.HexColor("#1B5E3B"),
        spaceAfter=8,
    )
    story = []
    story.append(
        Paragraph(
            f"<b>RESUMEN EJECUTIVO — {negocio.upper()}</b>",
            titulo_estilo,
        )
    )
    fecha_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    story.append(
        Paragraph(
            f"<b>Semana:</b> {semana} &nbsp;|&nbsp; <b>Emitido:</b> {fecha_str}",
            styles["Normal"],
        )
    )
    story.append(Spacer(1, 12))

    indicadores = [
        ["Indicador de control", "Valor registrado"],
        ["Semana de trabajo", semana],
        ["Total de despachos", str(n_despachos)],
        ["Total sacos despachados", str(total_sacos)],
        ["Facturación total (S/)", f"{total_monto:.2f}"],
        ["Entregas fiadas / por cobrar", str(n_fiados)],
        ["Monto por cobrar (S/)", f"{monto_fiado:.2f}"],
        ["Pollerías sin pedido", str(len(faltantes))],
    ]
    t = Table(indicadores, colWidths=[240, 250])
    t.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (1, 0), colors.HexColor("#1B5E3B")),
                ("TEXTCOLOR", (0, 0), (1, 0), colors.whitesmoke),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#F2F2F2")),
                ("GRID", (0, 0), (-1, -1), 1, colors.HexColor("#D9D9D9")),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
                ("FONTSIZE", (0, 0), (-1, -1), 10),
            ]
        )
    )
    story.append(t)
    story.append(Spacer(1, 14))

    if faltantes:
        story.append(
            Paragraph(
                "<b>Alerta — pollerías habituales sin despacho esta semana:</b>",
                styles["Normal"],
            )
        )
        story.append(Spacer(1, 4))
        for c in faltantes:
            story.append(Paragraph(f"• {c}", styles["Normal"]))
        story.append(Spacer(1, 10))
    else:
        story.append(
            Paragraph(
                "<b>Estado:</b> todas las pollerías de la lista tienen al menos un despacho.",
                styles["Normal"],
            )
        )
        story.append(Spacer(1, 10))

    # Tabla de detalle (máx. filas legibles)
    story.append(Paragraph("<b>Detalle de despachos</b>", styles["Heading2"]))
    story.append(Spacer(1, 6))
    if df.empty:
        story.append(Paragraph("Sin despachos registrados.", styles["Normal"]))
    else:
        head = ["Día", "Cliente", "Sacos", "Monto S/", "Estado"]
        data = [head]
        for _, row in df.iterrows():
            data.append(
                [
                    str(row["Día"]),
                    str(row["Cliente"])[:28],
                    str(int(row["Sacos"])),
                    f"{float(row['Monto S/']):.2f}",
                    str(row["Estado"])[:18],
                ]
            )
        # totales
        data.append(
            ["TOTAL", "", str(total_sacos), f"{total_monto:.2f}", ""]
        )
        det = Table(data, colWidths=[70, 160, 50, 70, 100])
        det.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
                    ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#1B5E3B")),
                    ("TEXTCOLOR", (0, -1), (-1, -1), colors.whitesmoke),
                    ("FONTSIZE", (0, 0), (-1, -1), 8),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CCCCCC")),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -2), [colors.white, colors.HexColor("#E8F0F8")]),
                    ("ALIGN", (2, 0), (3, -1), "RIGHT"),
                    ("TOPPADDING", (0, 0), (-1, -1), 3),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
                ]
            )
        )
        story.append(det)

    story.append(Spacer(1, 16))
    nota = Paragraph(
        "<i>Documento generado por Control de Papas y Pollerías. "
        "Los datos de este resumen corresponden al cuaderno semanal registrado en la app. "
        "Conservar junto al Excel del cuaderno para control de cobranza.</i>",
        styles["Normal"],
    )
    caja = Table([[nota]], colWidths=[500])
    caja.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#EDF2F7")),
                ("BOX", (0, 0), (-1, -1), 1, colors.HexColor("#1B5E3B")),
                ("PADDING", (0, 0), (-1, -1), 8),
            ]
        )
    )
    story.append(caja)

    doc.build(story)
    buffer.seek(0)
    return buffer
